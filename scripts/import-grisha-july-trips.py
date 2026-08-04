# -*- coding: utf-8 -*-
"""Activate Holvin account and backfill his missing July trips from registry."""
from __future__ import annotations

import json
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE = "https://armen4ik15-creator-transport-app-server-26b3.twc1.net"
ENV = Path(__file__).resolve().parent.parent / "timeweb.env"
REGISTRY = Path(r"C:\Users\Windows\Downloads\Telegram Desktop\Реестр июль 31.07.2026.xlsx")

DRIVER_ID = 16
CONTRACTOR_ID = 6  # ГК Стройавангард
LOGIN_EMAIL = "79515973189@reestrpro.local"
PHONE = "+7 (951) 597-31-89"
PASSWORD = "У596ЕА550"
CAR = "У596ЕА550"

vars_: dict[str, str] = {}
for line in ENV.read_text(encoding="utf-8").splitlines():
    if not line.strip() or line.strip().startswith("#") or "=" not in line:
        continue
    k, v = line.split("=", 1)
    vars_[k.strip()] = v.strip().strip('"').strip("'")

TOKEN = ""


def req(method: str, path: str, body: dict | None = None, retries: int = 6):
    data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    last_error: Exception | None = None
    for attempt in range(retries):
        request = urllib.request.Request(
            BASE + path,
            data=data,
            method=method,
            headers={
                "Authorization": f"Bearer {TOKEN}",
                "Content-Type": "application/json",
                "Accept": "application/json",
            },
        )
        try:
            with urllib.request.urlopen(request, timeout=120) as response:
                raw = response.read().decode("utf-8")
                return response.status, json.loads(raw) if raw else None
        except urllib.error.HTTPError as error:
            raw = error.read().decode("utf-8", "replace")
            try:
                parsed = json.loads(raw)
            except json.JSONDecodeError:
                parsed = {"error": raw[:500]}
            return error.code, parsed
        except Exception as error:  # noqa: BLE001
            last_error = error
            time.sleep(2 + attempt * 2)
    raise RuntimeError(f"request failed {method} {path}: {last_error}")


def login_admin() -> None:
    global TOKEN
    status, data = req(
        "POST",
        "/api/auth/login",
        {
            "email": vars_["FOUNDER_ADMIN_EMAIL"],
            "password": vars_["FOUNDER_ADMIN_PASSWORD"],
        },
    )
    if status >= 400:
        raise RuntimeError(f"admin login failed: {status} {data}")
    TOKEN = data["token"]
    print("admin login ok")


def load_grisha_rows() -> list[dict]:
    wb = openpyxl.load_workbook(REGISTRY, data_only=True)
    ws = wb["Лист1"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        item = dict(zip(headers, row))
        name = str(item.get("ФИО водителя") or "").upper()
        if "ХОЛВИН" not in name:
            continue
        dt = item.get("Дата")
        if isinstance(dt, datetime):
            item["Дата"] = dt.date().isoformat()
        rows.append(item)
    return rows


def order_key(row: dict) -> tuple:
    return (
        str(row.get("Материал") or "").strip(),
        str(row.get("Контрагент") or "").strip(),
        str(row.get("Погрузка") or "").strip(),
        str(row.get("Выгрузка") or "").strip().replace("\n", " "),
        row.get("Плечо"),
        row.get("Ставка водителя"),
        row.get("Ставка за м3-т"),
        str(row.get("Единица Измерения") or "").strip().lower().replace("м3", "м3"),
    )


def normalize_unit(raw: str) -> str:
    u = str(raw or "").strip().lower()
    if u in {"м3", "m3", "куб", "м³"}:
        return "м3"
    if u in {"т", "t", "тонн", "тонна"}:
        return "т"
    return str(raw or "м3").strip() or "м3"


def trip_timestamp(date_iso: str, index_on_day: int) -> str:
    # Stagger same-day trips: 10:15, 13:20, 16:40 Moscow-ish as UTC+3 -> store as ISO Z with local wall clock intent
    hours = [10, 13, 16, 18]
    minutes = [15, 20, 40, 5]
    h = hours[index_on_day % len(hours)]
    m = minutes[index_on_day % len(minutes)]
    # Store as plain timestamptz-looking string without forcing UTC shift of the calendar day
    return f"{date_iso} {h:02d}:{m:02d}:00+03"


def activate_driver() -> None:
    status, data = req(
        "PUT",
        f"/api/drivers/{DRIVER_ID}",
        {
            "email": LOGIN_EMAIL,
            "full_name": "ХОЛВИН ГРИГОРИЙ НИКОЛАЕВИЧ",
            "phone": PHONE,
            "car_number": CAR,
            "password": PASSWORD,
            "is_active": True,
        },
    )
    if status >= 400:
        raise RuntimeError(f"activate driver failed: {status} {data}")
    print("driver activated:", data.get("email"), data.get("phone"), data.get("car_number"), data.get("is_active"))


def ensure_orders(groups: dict[tuple, list[dict]]) -> dict[tuple, int]:
    status, orders = req("GET", "/api/orders")
    if status >= 400:
        raise RuntimeError(f"orders list failed: {status} {orders}")
    existing = [o for o in (orders or []) if int(o.get("driver_id") or 0) == DRIVER_ID]
    mapping: dict[tuple, int] = {}

    for key, trips in groups.items():
        material, _customer, load, unload, km, drv_rate, company_rate, unit = key
        if drv_rate is None or company_rate is None:
            print("SKIP order group (no rates):", material, "trips", len(trips))
            continue

        found = None
        for o in existing:
            if (
                str(o.get("material") or "").strip() == material
                and str(o.get("load_address") or "").strip() == load
                and str(o.get("unload_address") or "").strip().replace("\n", " ") == unload
                and float(o.get("driver_rate") or 0) == float(drv_rate)
                and float(o.get("company_rate") or 0) == float(company_rate)
            ):
                found = o
                break
        if found:
            mapping[key] = int(found["id"])
            print("reuse order", found["id"], material)
            continue

        body = {
            "driver_id": DRIVER_ID,
            "contractor_id": CONTRACTOR_ID,
            "task_name": f"{material} / {unload[:40]}",
            "material": material,
            "load_address": load,
            "unload_address": unload,
            "distance_km": float(km) if km is not None else None,
            "unit": normalize_unit(unit),
            "driver_rate": float(drv_rate),
            "company_rate": float(company_rate),
            "is_active": True,
            "status": "pending",
        }
        status, created = req("POST", "/api/orders", body)
        if status >= 400:
            raise RuntimeError(f"create order failed: {status} {created}")
        mapping[key] = int(created["id"])
        existing.append(created)
        print("created order", created["id"], material)
    return mapping


def main() -> None:
    login_admin()
    activate_driver()

    rows = load_grisha_rows()
    print("grisha registry rows", len(rows))

    groups: dict[tuple, list[dict]] = defaultdict(list)
    for row in rows:
        groups[order_key(row)].append(row)

    order_ids = ensure_orders(groups)

    # Existing TTNs in July window (and a bit wider)
    status, app_trips = req("GET", "/api/trips?from=2026-07-01&to=2026-08-05")
    if status >= 400:
        raise RuntimeError(f"trips list failed: {status} {app_trips}")
    existing_ttn = {
        str(t.get("ttn_number") or "").strip()
        for t in (app_trips or [])
        if t.get("ttn_number")
    }

    payload_trips = []
    skipped_incomplete = []
    already = []
    day_counters: dict[str, int] = defaultdict(int)

    for key, group_rows in groups.items():
        order_id = order_ids.get(key)
        if not order_id:
            for row in group_rows:
                skipped_incomplete.append(str(row.get("Номер ТН")))
            continue
        # sort by date then ttn for stable times
        ordered = sorted(
            group_rows,
            key=lambda r: (str(r.get("Дата")), str(r.get("Номер ТН"))),
        )
        for row in ordered:
            ttn = str(row.get("Номер ТН") or "").strip()
            if not ttn:
                continue
            if ttn in existing_ttn:
                already.append(ttn)
                continue
            date_iso = str(row.get("Дата"))
            idx = day_counters[date_iso]
            day_counters[date_iso] += 1
            vol = row.get("всего")
            try:
                volume = float(vol) if vol is not None and str(vol).strip() != "" else None
            except (TypeError, ValueError):
                volume = None
            payload_trips.append(
                {
                    "order_id": order_id,
                    "ttn_number": ttn,
                    "volume": volume,
                    "trip_at": trip_timestamp(date_iso, idx),
                    "note": "imported from july registry (no photo yet)",
                }
            )

    print("to create", len(payload_trips), "already", len(already), "incomplete", skipped_incomplete)

    if not payload_trips:
        print("nothing to import")
        return

    status, result = req("POST", "/api/trips/backfill", {"trips": payload_trips})
    print("backfill status", status)
    print(json.dumps(result, ensure_ascii=False, indent=2)[:4000])
    if status >= 400:
        raise SystemExit(1)

    # Verify login as driver by phone
    status, driver_login = req(
        "POST",
        "/api/auth/login",
        {"email": "79515973189", "password": PASSWORD},
    )
    print("driver login by phone digits:", status, (driver_login or {}).get("user", {}).get("full_name"))
    status2, driver_login2 = req(
        "POST",
        "/api/auth/login",
        {"email": LOGIN_EMAIL, "password": PASSWORD},
    )
    print("driver login by email:", status2, (driver_login2 or {}).get("user", {}).get("email"))


if __name__ == "__main__":
    main()
